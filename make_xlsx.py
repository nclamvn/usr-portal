#!/usr/bin/env python3
"""Export uav-300.xlsx from out/site-data.json (the canonical REAL set, now 302) — NOT from bundle.html
(which carries pre-rendered fragments, no #site-data JSON block). Same honest-null contract as the
site: a field with no verified value -> EMPTY cell (never 0 / "N/A" / "–"). Disputed values are not
collapsed into sheet 1; both claims live in the 'Tranh chấp' sheet. The 'Tổng (sống)' sheet uses
Excel FORMULAS over sheet 1 so totals recompute live — nothing hand-typed.
"""
import json, pathlib
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

ROOT = pathlib.Path(__file__).resolve().parent
SITE = ROOT / "out" / "site-data.json"
OUT = ROOT / "uav-300.xlsx"

# (key, header, is_top_level_string?) — column order for sheet 1
COLS = [
    ("slug", "slug", "top"),
    ("manufacturer", "Hãng", "cell"),
    ("name", "Mẫu", "cell"),
    ("manufacturer_country", "Quốc gia", "cell"),
    ("market_segment", "Phân khúc", "cell"),
    ("provenance_class", "Lớp", "top"),
    ("airframe_type", "Khung bay", "cell"),
    ("propulsion", "Động lực", "cell"),
    ("live_status", "Tình trạng", "cell"),
    ("mtow_kg", "MTOW (kg)", "cell"),
    ("max_payload_kg", "Tải tối đa (kg)", "cell"),
    ("endurance_min", "Thời gian bay (phút)", "cell"),
    ("max_range_km", "Tầm bay (km)", "cell"),
    ("max_link_km", "Tầm truyền (km)", "cell"),
    ("max_speed_ms", "Tốc độ tối đa (m/s)", "cell"),
    ("service_ceiling_m", "Trần bay (m)", "cell"),
    ("encryption", "Mã hoá", "cell"),
    ("blue_uas", "Blue UAS", "cell"),
    ("ndaa_compliant", "NDAA", "cell"),
    ("gps_denied_capable", "GPS-denied", "cell"),
]
SPEC_KEYS = ["mtow_kg", "max_payload_kg", "endurance_min", "max_range_km", "max_link_km",
             "max_speed_ms", "service_ceiling_m", "encryption", "blue_uas", "ndaa_compliant",
             "gps_denied_capable"]
HEAD = Font(bold=True)
MONO = Font(name="Consolas")


def cell_value(e, key, kind):
    """Honest-null: only a real (non-null) value reaches the sheet; otherwise None -> empty cell.
    field_obj already nulls 'unverified' and 'disputed', so e[key]['value'] is the honest value."""
    if kind == "top":
        return e.get(key)
    return e.get(key, {}).get("value")   # None when unverified/absent/disputed -> blank


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        ws.cell(row=1, column=c).font = HEAD
    ws.freeze_panes = "A2"


def main():
    site = json.loads(SITE.read_bytes())
    ents = sorted((e for e in site["entities"] if e.get("entity_type", "uav") == "uav"),
                  key=lambda e: e["slug"])  # schema/2: UAV rows only
    all_fields = site["field_groups"]["display"] + site["field_groups"]["spec"]
    wb = Workbook()

    # ---- Sheet 1: UAV (200) ----
    ws = wb.active
    ws.title = "UAV (200)"
    ws.append([h for _, h, _ in COLS])
    for e in ents:
        ws.append([cell_value(e, k, kind) for k, _, kind in COLS])
    style_header(ws, len(COLS))
    ws.column_dimensions["A"].width = 42
    for c in range(2, len(COLS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    # ---- Sheet 2: Nguồn & tier ----
    ws2 = wb.create_sheet("Nguồn & tier")
    ws2.append(["slug", "Trường", "Giá trị", "Tier", "URL", "Ngày lấy"])
    nsrc = 0
    for e in ents:
        for fkey in all_fields:
            fo = e.get(fkey, {})
            for s in fo.get("sources", []) or []:
                ws2.append([e["slug"], fkey, fo.get("value"), s.get("tier"),
                            s.get("url"), s.get("retrieved")])
                nsrc += 1
    style_header(ws2, 6)
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["E"].width = 60

    # ---- Sheet 3: Tranh chấp (both claims, never resolved) ----
    ws3 = wb.create_sheet("Tranh chấp")
    ws3.append(["slug", "Trường", "Giá trị claim", "Tier", "URL"])
    ndisp = nclaim = 0
    for e in ents:
        for fkey in all_fields:
            fo = e.get(fkey, {})
            if fo.get("status") == "disputed":
                ndisp += 1
                for c in fo.get("claims", []) or []:
                    ws3.append([e["slug"], fkey, c.get("claimed_value"), c.get("tier"), c.get("url")])
                    nclaim += 1
    style_header(ws3, 5)
    ws3.column_dimensions["A"].width = 42
    ws3.column_dimensions["E"].width = 60

    # ---- Sheet 4: Tổng (sống) — Excel formulas over sheet 1, recompute live ----
    ws4 = wb.create_sheet("Tổng (sống)")
    s1 = "'UAV (200)'"
    n = len(ents)
    last = n + 1                                    # data rows 2..last
    col = {k: get_column_letter(i + 1) for i, (k, _, _) in enumerate(COLS)}
    blue, ndaa, gps = col["blue_uas"], col["ndaa_compliant"], col["gps_denied_capable"]
    spec_first = col[SPEC_KEYS[0]]                  # mtow_kg
    spec_last = col[SPEC_KEYS[-1]]                  # gps_denied_capable
    ws4.append(["Chỉ số", "Giá trị (công thức sống)"])
    rows = [
        ("Tổng thực thể", f"=COUNTA({s1}!A2:A{last})"),
        ("Blue UAS (verified)", f"=COUNTIF({s1}!{blue}2:{blue}{last},TRUE)"),
        ("NDAA (verified)", f"=COUNTIF({s1}!{ndaa}2:{ndaa}{last},TRUE)"),
        ("GPS-denied (verified)", f"=COUNTIF({s1}!{gps}2:{gps}{last},TRUE)"),
        ("Độ phủ spec", f"=COUNTA({s1}!{spec_first}2:{spec_last}{last})/({n}*{len(SPEC_KEYS)})"),
        ("Số claim tranh chấp", "=COUNTA('Tranh chấp'!A2:A10000)"),
    ]
    for label, formula in rows:
        ws4.append([label, formula])
    ws4.cell(row=6, column=2).number_format = "0%"   # coverage as %
    style_header(ws4, 2)
    ws4.column_dimensions["A"].width = 26
    ws4.column_dimensions["B"].width = 26

    wb.save(OUT)
    print(f"uav-300.xlsx: {n} entities · {nsrc} source rows · {ndisp} disputed cells/{nclaim} claims")


if __name__ == "__main__":
    main()
